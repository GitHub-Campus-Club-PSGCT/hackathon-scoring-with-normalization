import io
import json
import os
import uuid
from datetime import datetime

import pandas as pd
from flask import Flask, Response, flash, redirect, render_template, request, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import app as scoring_app


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "thooral-bulk-secret-key-it-is")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
BULK_JOBS_DIR = os.path.join(DATA_DIR, "bulk_jobs")

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_ERROR_SAMPLES = 25


def ensure_dirs():
    os.makedirs(BULK_JOBS_DIR, exist_ok=True)


def allowed_file(filename):
    return os.path.splitext(filename.lower())[1] in ALLOWED_EXTENSIONS


def create_job_dir():
    ensure_dirs()
    job_id = uuid.uuid4().hex
    job_dir = os.path.join(BULK_JOBS_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)
    return job_id, job_dir


def get_job_dir(job_id):
    return os.path.join(BULK_JOBS_DIR, job_id)


def get_meta_path(job_id):
    return os.path.join(get_job_dir(job_id), "meta.json")


def get_data_path(job_id):
    return os.path.join(get_job_dir(job_id), "input.csv")


def get_results_path(job_id):
    return os.path.join(get_job_dir(job_id), "results.json")


def load_df(job_id):
    data_path = get_data_path(job_id)
    if not os.path.exists(data_path):
        return None
    return pd.read_csv(data_path, dtype=object).fillna("")


def save_meta(job_id, payload):
    with open(get_meta_path(job_id), "w") as meta_file:
        json.dump(payload, meta_file, indent=2)


def load_meta(job_id):
    meta_path = get_meta_path(job_id)
    if not os.path.exists(meta_path):
        return None
    with open(meta_path, "r") as meta_file:
        return json.load(meta_file)


def read_uploaded_dataframe(upload):
    ext = os.path.splitext(upload.filename.lower())[1]
    if ext == ".csv":
        df = pd.read_csv(upload, dtype=object)
    else:
        df = pd.read_excel(upload, dtype=object)
    return df.fillna("")


def normalize_with_existing_logic(all_scores, criteria):
    original_get_all_scores = scoring_app.get_all_scores
    original_get_criteria = scoring_app.get_criteria
    try:
        scoring_app.get_all_scores = lambda round_id=None: all_scores
        scoring_app.get_criteria = lambda: criteria
        return scoring_app.normalize_scores(None)
    finally:
        scoring_app.get_all_scores = original_get_all_scores
        scoring_app.get_criteria = original_get_criteria


def build_numeric_validation_errors(df, score_fields):
    errors = []
    for field in score_fields:
        series = df[field]
        numeric_series = pd.to_numeric(series, errors="coerce")
        original_as_text = series.astype(str).str.strip()
        invalid_mask = original_as_text.ne("") & numeric_series.isna()
        invalid_rows = df[invalid_mask]
        for row_idx, row in invalid_rows.head(MAX_ERROR_SAMPLES).iterrows():
            errors.append(
                {
                    "row_number": int(row_idx) + 2,
                    "field": field,
                    "value": str(row[field]),
                }
            )
            if len(errors) >= MAX_ERROR_SAMPLES:
                return errors
    return errors


def build_all_scores_payload(df, evaluator_field, entry_field, score_fields):
    all_scores = []
    criteria = []

    for field in score_fields:
        numeric_values = pd.to_numeric(df[field], errors="coerce")
        if numeric_values.dropna().empty:
            min_score = 0.0
            max_score = 0.0
        else:
            min_score = float(numeric_values.min())
            max_score = float(numeric_values.max())
        criteria.append(
            {
                "id": field,
                "name": field,
                "type": "adder",
                "min_score": min_score,
                "max_score": max_score,
            }
        )

    numeric_df = df.copy()
    for field in score_fields:
        numeric_df[field] = pd.to_numeric(numeric_df[field], errors="coerce").fillna(0.0)

    for _, row in numeric_df.iterrows():
        evaluator = str(row[evaluator_field]).strip()
        entry_id = str(row[entry_field]).strip()
        record = {
            "judge": evaluator,
            "team_id": entry_id,
            "team_name": entry_id,
        }
        for field in score_fields:
            record[field] = float(row[field])
        all_scores.append(record)

    return all_scores, criteria


@app.route("/", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        upload_file = request.files.get("file")
        if not upload_file or not upload_file.filename:
            flash("Please choose a CSV or Excel file.", "error")
            return redirect(url_for("upload"))

        if not allowed_file(upload_file.filename):
            flash("Unsupported file format. Upload CSV, XLSX, or XLS only.", "error")
            return redirect(url_for("upload"))

        try:
            df = read_uploaded_dataframe(upload_file)
        except Exception as exc:
            flash(f"Could not read file: {exc}", "error")
            return redirect(url_for("upload"))

        if df.empty:
            flash("The uploaded file is empty.", "error")
            return redirect(url_for("upload"))

        columns = [str(col) for col in df.columns]
        job_id, _ = create_job_dir()

        data_path = get_data_path(job_id)
        df.to_csv(data_path, index=False)
        save_meta(
            job_id,
            {
                "filename": upload_file.filename,
                "columns": columns,
                "created_at": datetime.now().isoformat(),
            },
        )

        return redirect(url_for("configure", job_id=job_id))

    return render_template("bulk_upload.html")


@app.route("/configure/<job_id>", methods=["GET", "POST"])
def configure(job_id):
    meta = load_meta(job_id)
    df = load_df(job_id)
    if not meta or df is None:
        flash("Job not found or expired.", "error")
        return redirect(url_for("upload"))

    columns = meta.get("columns", [])
    selected = meta.get("selected", {})

    if request.method == "POST":
        evaluator_field = request.form.get("evaluator_field", "").strip()
        entry_field = request.form.get("entry_field", "").strip()
        score_fields = request.form.getlist("score_fields")

        if not evaluator_field or evaluator_field not in columns:
            flash("Please select a valid evaluator field.", "error")
            return render_template(
                "bulk_configure.html",
                job_id=job_id,
                columns=columns,
                selected={
                    "evaluator_field": evaluator_field,
                    "entry_field": entry_field,
                    "score_fields": score_fields,
                },
            )

        if not entry_field or entry_field not in columns:
            flash("Please select a valid team ID field.", "error")
            return render_template(
                "bulk_configure.html",
                job_id=job_id,
                columns=columns,
                selected={
                    "evaluator_field": evaluator_field,
                    "entry_field": entry_field,
                    "score_fields": score_fields,
                },
            )

        score_fields = [field for field in score_fields if field in columns]
        if not score_fields:
            flash("Select at least one scoring criteria field.", "error")
            return render_template(
                "bulk_configure.html",
                job_id=job_id,
                columns=columns,
                selected={
                    "evaluator_field": evaluator_field,
                    "entry_field": entry_field,
                    "score_fields": score_fields,
                },
            )

        if evaluator_field in score_fields or entry_field in score_fields:
            flash("Evaluator and entry fields cannot be selected as scoring criteria.", "error")
            return render_template(
                "bulk_configure.html",
                job_id=job_id,
                columns=columns,
                selected={
                    "evaluator_field": evaluator_field,
                    "entry_field": entry_field,
                    "score_fields": score_fields,
                },
            )

        if evaluator_field == entry_field:
            flash("Evaluator field and team ID field must be different.", "error")
            return render_template(
                "bulk_configure.html",
                job_id=job_id,
                columns=columns,
                selected={
                    "evaluator_field": evaluator_field,
                    "entry_field": entry_field,
                    "score_fields": score_fields,
                },
            )

        validation_errors = build_numeric_validation_errors(df, score_fields)
        if validation_errors:
            return render_template(
                "bulk_configure.html",
                job_id=job_id,
                columns=columns,
                selected={
                    "evaluator_field": evaluator_field,
                    "entry_field": entry_field,
                    "score_fields": score_fields,
                },
                validation_errors=validation_errors,
            )

        evaluator_counts_df = (
            df.groupby(evaluator_field, dropna=False).size().reset_index(name="count")
        )
        evaluator_counts = []
        for _, row in evaluator_counts_df.iterrows():
            evaluator_counts.append(
                {
                    "evaluator": str(row[evaluator_field]).strip(),
                    "count": int(row["count"]),
                }
            )

        selected = {
            "evaluator_field": evaluator_field,
            "entry_field": entry_field,
            "score_fields": score_fields,
            "total_rows": int(len(df)),
            "evaluator_counts": evaluator_counts,
        }
        meta["selected"] = selected
        save_meta(job_id, meta)

        return render_template(
            "bulk_confirm.html",
            job_id=job_id,
            selected=selected,
            filename=meta.get("filename", "uploaded file"),
        )

    return render_template(
        "bulk_configure.html",
        job_id=job_id,
        columns=columns,
        selected=selected,
    )


@app.route("/confirm/<job_id>", methods=["POST"])
def confirm(job_id):
    meta = load_meta(job_id)
    df = load_df(job_id)
    if not meta or df is None:
        flash("Job not found or expired.", "error")
        return redirect(url_for("upload"))

    selected = meta.get("selected")
    if not selected:
        flash("Please configure evaluator and scoring fields first.", "error")
        return redirect(url_for("configure", job_id=job_id))

    evaluator_field = selected["evaluator_field"]
    entry_field = selected["entry_field"]
    score_fields = selected["score_fields"]

    all_scores, criteria = build_all_scores_payload(df, evaluator_field, entry_field, score_fields)
    results = normalize_with_existing_logic(all_scores, criteria)

    with open(get_results_path(job_id), "w") as results_file:
        json.dump(results, results_file, indent=2)

    return redirect(url_for("results", job_id=job_id))


@app.route("/results/<job_id>")
def results(job_id):
    meta = load_meta(job_id)
    results_path = get_results_path(job_id)
    if not meta or not os.path.exists(results_path):
        flash("No results available for this job.", "error")
        return redirect(url_for("upload"))

    with open(results_path, "r") as results_file:
        team_results = json.load(results_file)

    return render_template(
        "bulk_results.html",
        job_id=job_id,
        results=team_results,
        selected=meta.get("selected", {}),
        filename=meta.get("filename", "uploaded file"),
    )


@app.route("/download/<job_id>")
def download(job_id):
    meta = load_meta(job_id)
    df = load_df(job_id)
    results_path = get_results_path(job_id)
    if not meta or df is None or not os.path.exists(results_path):
        flash("No downloadable results found for this job.", "error")
        return redirect(url_for("upload"))

    with open(results_path, "r") as results_file:
        team_results = json.load(results_file)

    selected = meta.get("selected", {})
    evaluator_field = selected.get("evaluator_field", "")
    entry_field = selected.get("entry_field", "")
    score_fields = selected.get("score_fields", [])

    wb = Workbook()
    ws_rankings = wb.active
    ws_rankings.title = "Rankings"
    ws_raw = wb.create_sheet(title="Raw Scores")

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ranking_headers = [
        "Rank",
        "Team ID",
        "Team Name",
        "Avg Raw Score",
        "Avg Normalized Score",
        "Num Evaluators",
    ]
    for col, text in enumerate(ranking_headers, start=1):
        cell = ws_rankings.cell(row=1, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(team_results, start=2):
        ws_rankings.cell(row=idx, column=1, value=idx - 1).border = thin_border
        ws_rankings.cell(row=idx, column=2, value=row.get("team_id", "")).border = thin_border
        ws_rankings.cell(row=idx, column=3, value=row.get("team_name", "")).border = thin_border
        ws_rankings.cell(row=idx, column=4, value=row.get("avg_raw_score", 0)).border = thin_border
        ws_rankings.cell(row=idx, column=5, value=row.get("avg_normalized_score", 0)).border = thin_border
        ws_rankings.cell(row=idx, column=6, value=row.get("num_judges", 0)).border = thin_border

    ws_rankings.column_dimensions["A"].width = 8
    ws_rankings.column_dimensions["B"].width = 24
    ws_rankings.column_dimensions["C"].width = 24
    ws_rankings.column_dimensions["D"].width = 18
    ws_rankings.column_dimensions["E"].width = 22
    ws_rankings.column_dimensions["F"].width = 16

    raw_headers = [evaluator_field, entry_field] + score_fields
    for col, text in enumerate(raw_headers, start=1):
        cell = ws_raw.cell(row=1, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ordered_values = [row.get(evaluator_field, ""), row.get(entry_field, "")]
        for field in score_fields:
            ordered_values.append(row.get(field, ""))
        for col_idx, value in enumerate(ordered_values, start=1):
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 2:
                cell.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_prefix = os.path.splitext(meta.get("filename", "scores"))[0]
    out_name = f"{filename_prefix}_normalized_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={out_name}"},
    )


if __name__ == "__main__":
    ensure_dirs()
    app.run(host="0.0.0.0", port=6061, debug=True)
