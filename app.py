import os
import json
import csv
from datetime import datetime
from functools import wraps
from filelock import FileLock
import numpy as np
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.middleware.dispatcher import DispatcherMiddleware
from werkzeug.serving import run_simple

app = Flask(__name__)
app.secret_key = os.urandom(24)

# File paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')
SCORES_FILE = os.path.join(BASE_DIR, 'scores.csv')
LOCK_FILE = os.path.join(BASE_DIR, 'scores.csv.lock')


def load_config():
    """Load configuration from JSON file."""
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)


def get_judges():
    """Get list of judges from config."""
    config = load_config()
    return {j['username']: j['password'] for j in config['judges']}


def get_admins():
    """Get list of admins from config."""
    config = load_config()
    return {a['username']: a['password'] for a in config.get('admins', [])}


def get_teams():
    """Get list of teams from config."""
    config = load_config()
    return config['teams']


def get_criteria():
    """Get judging criteria from config."""
    config = load_config()
    return config['criteria']


def init_scores_file():
    """Initialize the scores CSV file if it doesn't exist."""
    if not os.path.exists(SCORES_FILE):
        criteria = get_criteria()
        headers = ['timestamp', 'judge', 'team_id', 'team_name'] + [c['id'] for c in criteria]
        with open(SCORES_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)


def save_scores(judge, team_id, team_name, scores):
    """Save scores to CSV file with file locking for concurrent access."""
    lock = FileLock(LOCK_FILE, timeout=10)
    
    with lock:
        # Read existing scores
        existing_rows = []
        if os.path.exists(SCORES_FILE):
            with open(SCORES_FILE, 'r', newline='') as f:
                reader = csv.reader(f)
                existing_rows = list(reader)
        
        # Get headers
        criteria = get_criteria()
        headers = ['timestamp', 'judge', 'team_id', 'team_name'] + [c['id'] for c in criteria]
        
        # If file is empty or doesn't have headers, add them
        if not existing_rows:
            existing_rows = [headers]
        
        # Check if this judge already scored this team
        updated = False
        for i, row in enumerate(existing_rows[1:], start=1):
            if len(row) >= 3 and row[1] == judge and row[2] == team_id:
                # Update existing score
                timestamp = datetime.now().isoformat()
                new_row = [timestamp, judge, team_id, team_name] + [scores.get(c['id'], 0) for c in criteria]
                existing_rows[i] = new_row
                updated = True
                break
        
        # If no existing score found, append new one
        if not updated:
            timestamp = datetime.now().isoformat()
            new_row = [timestamp, judge, team_id, team_name] + [scores.get(c['id'], 0) for c in criteria]
            existing_rows.append(new_row)
        
        # Write back all rows
        with open(SCORES_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(existing_rows)


def get_all_scores():
    """Read all scores from CSV file."""
    if not os.path.exists(SCORES_FILE):
        return []
    
    lock = FileLock(LOCK_FILE, timeout=10)
    with lock:
        with open(SCORES_FILE, 'r', newline='') as f:
            reader = csv.DictReader(f)
            return list(reader)


def get_judge_scores(judge):
    """Get all scores submitted by a specific judge."""
    all_scores = get_all_scores()
    return [s for s in all_scores if s.get('judge') == judge]


def get_team_score_by_judge(judge, team_id):
    """Get scores for a specific team by a specific judge."""
    judge_scores = get_judge_scores(judge)
    for score in judge_scores:
        if score.get('team_id') == team_id:
            return score
    return None


def normalize_scores():
    """
    Normalize scores using Z-score normalization per judge.
    This accounts for judges who are strict or lenient.
    Each judge's scores are normalized to have mean=0, std=1,
    then transformed to a 0-100 scale.
    """
    all_scores = get_all_scores()
    if not all_scores:
        return []
    
    criteria = get_criteria()
    criteria_ids = [c['id'] for c in criteria]
    
    # Convert to DataFrame for easier manipulation
    df = pd.DataFrame(all_scores)
    
    if df.empty:
        return []
    
    # Convert score columns to numeric
    for cid in criteria_ids:
        if cid in df.columns:
            df[cid] = pd.to_numeric(df[cid], errors='coerce').fillna(0)
    
    # Calculate total raw score for each entry
    df['total_raw'] = df[criteria_ids].sum(axis=1)
    
    # Z-score normalize per judge
    normalized_data = []
    
    for judge in df['judge'].unique():
        judge_df = df[df['judge'] == judge].copy()
        
        if len(judge_df) > 1:
            # Calculate mean and std for this judge's total scores
            mean_score = judge_df['total_raw'].mean()
            std_score = judge_df['total_raw'].std()
            
            if std_score > 0:
                # Z-score normalization
                judge_df['z_score'] = (judge_df['total_raw'] - mean_score) / std_score
            else:
                judge_df['z_score'] = 0
        else:
            # Only one score from this judge, can't normalize
            judge_df['z_score'] = 0
        
        normalized_data.append(judge_df)
    
    if not normalized_data:
        return []
    
    result_df = pd.concat(normalized_data, ignore_index=True)
    
    # Convert Z-scores to 0-100 scale (assuming Z-scores typically range from -3 to 3)
    # Map z-score to 0-100: z=-3 -> 0, z=0 -> 50, z=3 -> 100
    result_df['normalized_score'] = ((result_df['z_score'] + 3) / 6 * 100).clip(0, 100).round(2)
    
    # Aggregate by team
    team_results = []
    for team_id in result_df['team_id'].unique():
        team_data = result_df[result_df['team_id'] == team_id]
        team_name = team_data['team_name'].iloc[0]
        
        avg_raw = team_data['total_raw'].mean()
        avg_normalized = team_data['normalized_score'].mean()
        num_judges = len(team_data)
        
        # Individual judge scores for this team
        judge_scores = []
        for _, row in team_data.iterrows():
            judge_scores.append({
                'judge': row['judge'],
                'raw_score': row['total_raw'],
                'normalized_score': row['normalized_score']
            })
        
        team_results.append({
            'team_id': team_id,
            'team_name': team_name,
            'avg_raw_score': round(avg_raw, 2),
            'avg_normalized_score': round(avg_normalized, 2),
            'num_judges': num_judges,
            'judge_scores': judge_scores
        })
    
    # Sort by normalized score descending
    team_results.sort(key=lambda x: x['avg_normalized_score'], reverse=True)
    
    return team_results


def login_required(f):
    """Decorator to require login for a route."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'judge' not in session and 'admin' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Decorator to require admin login for a route."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin' not in session:
            flash('Admin access required', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# Initialize scores file on startup
init_scores_file()


@app.route('/')
def index():
    """Redirect to login or dashboard."""
    if 'admin' in session:
        return redirect(url_for('results'))
    if 'judge' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle judge or admin login."""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        # Check admin credentials first
        admins = get_admins()
        if username in admins and admins[username] == password:
            session['admin'] = username
            flash('Admin login successful', 'success')
            return redirect(url_for('results'))
        
        # Then check judge credentials
        judges = get_judges()
        if username in judges and judges[username] == password:
            session['judge'] = username
            flash('Login successful', 'success')
            return redirect(url_for('dashboard'))
        
        flash('Invalid username or password', 'error')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Handle judge or admin logout."""
    session.pop('judge', None)
    session.pop('admin', None)
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard showing teams to score."""
    judge = session['judge']
    teams = get_teams()
    judge_scores = get_judge_scores(judge)
    
    # Mark which teams have been scored
    scored_team_ids = {s['team_id'] for s in judge_scores}
    
    teams_with_status = []
    for team in teams:
        teams_with_status.append({
            **team,
            'scored': team['id'] in scored_team_ids
        })
    
    return render_template('dashboard.html', 
                         judge=judge, 
                         teams=teams_with_status,
                         scored_count=len(scored_team_ids),
                         total_count=len(teams))


@app.route('/score/<team_id>', methods=['GET', 'POST'])
@login_required
def score_team(team_id):
    """Score a specific team."""
    judge = session['judge']
    teams = get_teams()
    criteria = get_criteria()
    
    # Find the team
    team = None
    for t in teams:
        if t['id'] == team_id:
            team = t
            break
    
    if not team:
        flash('Team not found', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        scores = {}
        for c in criteria:
            score_value = request.form.get(c['id'], 0)
            try:
                score_value = int(score_value)
                score_value = max(0, min(score_value, c['max_score']))
            except (ValueError, TypeError):
                score_value = 0
            scores[c['id']] = score_value
        
        save_scores(judge, team['id'], team['name'], scores)
        flash(f'Scores saved for {team["name"]}', 'success')
        return redirect(url_for('dashboard'))
    
    # Get existing scores if any
    existing_scores = get_team_score_by_judge(judge, team_id)
    
    return render_template('score.html',
                         judge=judge,
                         team=team,
                         criteria=criteria,
                         existing_scores=existing_scores)


@app.route('/results')
@admin_required
def results():
    """View normalized results (admin only)."""
    admin = session['admin']
    team_results = normalize_scores()
    criteria = get_criteria()
    max_possible = sum(c['max_score'] for c in criteria)
    
    return render_template('results.html',
                         admin=admin,
                         results=team_results,
                         max_possible=max_possible)


@app.route('/export-results')
@admin_required
def export_results():
    """Export results to XLSX with two sheets: Rankings and Raw Scores."""
    from flask import Response
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    team_results = normalize_scores()
    criteria = get_criteria()
    max_possible = sum(c['max_score'] for c in criteria)
    all_scores = get_all_scores()
    
    # Create workbook
    wb = Workbook()
    
    # --- Sheet 1: Rankings ---
    ws1 = wb.active
    ws1.title = "Rankings"
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers for Rankings
    ranking_headers = ['Rank', 'Team ID', 'Team Name', 'Avg Raw Score', 'Max Possible', 'Avg Normalized Score', 'Num Judges']
    for col, header in enumerate(ranking_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Write ranking data
    for i, result in enumerate(team_results, 1):
        row_data = [
            i,
            result['team_id'],
            result['team_name'],
            result['avg_raw_score'],
            max_possible,
            result['avg_normalized_score'],
            result['num_judges']
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=i+1, column=col, value=value)
            cell.border = thin_border
            if col in [1, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths for Rankings
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 12
    
    # --- Sheet 2: Raw Scores (scores.csv data) ---
    ws2 = wb.create_sheet(title="Raw Scores")
    
    # Get headers from criteria
    criteria_ids = [c['id'] for c in criteria]
    criteria_names = {c['id']: c['name'] for c in criteria}
    raw_headers = ['Timestamp', 'Judge', 'Team ID', 'Team Name'] + [criteria_names.get(cid, cid) for cid in criteria_ids]
    
    # Write headers for Raw Scores
    for col, header in enumerate(raw_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Write raw score data
    for row_idx, score in enumerate(all_scores, 2):
        row_data = [
            score.get('timestamp', ''),
            score.get('judge', ''),
            score.get('team_id', ''),
            score.get('team_name', '')
        ] + [score.get(cid, 0) for cid in criteria_ids]
        
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col > 4:  # Score columns
                cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths for Raw Scores
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 25
    for i, _ in enumerate(criteria_ids, 5):
        ws2.column_dimensions[chr(64 + i)].width = 18
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=hackathon_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )


@app.route('/api/scores', methods=['GET'])
@admin_required
def api_scores():
    """API endpoint to get all scores (admin only)."""
    return jsonify(get_all_scores())


@app.route('/api/normalized', methods=['GET'])
@admin_required
def api_normalized():
    """API endpoint to get normalized scores (admin only)."""
    return jsonify(normalize_scores())


if __name__ == '__main__':
    # Create a dispatcher that mounts the app at /scoring
    application = DispatcherMiddleware(
        Flask(__name__),  # Dummy app for root
        {'/scoring': app}
    )
    run_simple('0.0.0.0', 6060, application, use_reloader=True, use_debugger=True)
